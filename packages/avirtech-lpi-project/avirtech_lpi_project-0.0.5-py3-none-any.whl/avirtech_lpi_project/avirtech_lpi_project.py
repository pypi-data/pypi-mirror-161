from email.mime import base
from tkinter.tix import DirList
from tkFileDialog import askopenfilename
import os

# library yang dibutuhkan dari kodingan statistik ilham
from msilib.schema import Error
import arcpy
import tkinter as tk
from tkinter import messagebox
import tkFileDialog as filedialog
import os
import openpyxl
import excel2img
from simpledbf import Dbf5
from simpledbf import Dbf5
from fpdf import FPDF
import pandas as pd
import matplotlib.pyplot as plt
from pandas.tools.plotting import table
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl.styles.alignment import Alignment
import os
import matplotlib.pyplot as plt
import pandas as pd
from os.path import exists

class weed_detection:
    @staticmethod
    def weed_detection():
        location = os.path.expanduser('~/Documents/Avirtech/Avirkey/Avirkey.ini')

        if exists(location):
            root = tk.Tk()
            root.withdraw()
            messagebox.showinfo("","Please input your raster data folder location")
            raster = filedialog.askdirectory()
            messagebox.showinfo("","Please input your shapefile data folder location")
            shapefile = filedialog.askdirectory()
            messagebox.showinfo("","Please input your .ecd folder location")
            model = filedialog.askdirectory()
            messagebox.showinfo("","Please input your .lyr symbology folder location")
            symbol = filedialog.askdirectory()
            messagebox.showinfo("","Please input your output folder destination")
            root_folder = filedialog.askdirectory()
            root.destroy

            dir_extract = os.path.join(root_folder, "1._extract")
            os.mkdir(dir_extract)
            dir_convert = os.path.join(root_folder, "2._convert")
            os.mkdir(dir_convert)
            dir_svm = os.path.join(root_folder, "3._classification")
            os.mkdir(dir_svm)

            mxd = arcpy.mapping.MapDocument("Current")
            arcpy.env.workspace = "Current"
            global df
            df = arcpy.mapping.ListDataFrames(mxd)[0]

            arcpy.CreateFileGDB_management(dir_extract, "Otomasi.gdb", "CURRENT")
            loc_gdb = os.path.join(dir_extract, "Otomasi.gdb")

            data_raster = []
            for file in os.listdir(raster):
                if file.endswith('.tif'):
                    arcpy.MakeRasterLayer_management(os.path.join(raster,file), file, "", os.path.join(raster,file))
                    data_raster.append(file)

            data_shapefile = []
            for file in os.listdir(shapefile):
                if file.endswith('.shp'):
                    base = os.path.splitext(file)[0]
                    Layer2 = arcpy.mapping.Layer(os.path.join(shapefile, file))
                    arcpy.mapping.AddLayer(df,Layer2,"BOTTOM")
                    data_shapefile.append(base)

            data_model = []
            for file in os.listdir(model):
                if file.endswith('.ecd',):
                    data_model.append(os.path.join(model, file))

            data_symbology = []
            for file in os.listdir(symbol):
                if file.endswith('.lyr'):
                    data_symbology.append(os.path.join(symbol, file))


            for shp in data_shapefile:
                list_petak = list([(row.getValue("Petak_12")) for row in arcpy.SearchCursor(shp, fields="Petak_12")])
                
                global i
                for i in list_petak:

                    arcpy.SelectLayerByAttribute_management(shp, "NEW_SELECTION", "\"Petak_12\" = '{}'".format(i))

                    arcpy.gp.ExtractByMask_sa(data_raster[0], shp, os.path.join(loc_gdb,"Petak_{}".format(i)))

                    loc_convert = os.path.join(dir_convert, "{}".format(i))

                    os.mkdir(loc_convert)

                    arcpy.RasterToOtherFormat_conversion(os.path.join(loc_gdb,"Petak_{}".format(i)), loc_convert, "TIFF")

                    loc_svm = os.path.join(dir_svm, "{}".format(i))
                    loc_polygon = os.path.join(loc_svm, "polygon")

                    os.mkdir(loc_svm)
                    os.mkdir(loc_polygon)


                    global nama_petak
                    nama_petak = os.path.basename(os.path.normpath(loc_svm))
                    nama_petak_low = nama_petak.lower()
                    
                    arcpy.gp.ClassifyRaster_sa(os.path.join(loc_convert,"Petak_{}.tif".format(i)), data_model[0], os.path.join(loc_svm, "SVM_{}.tif".format(i)), "")

                    arcpy.RasterToPolygon_conversion(os.path.join(loc_svm, "SVM_{}.tif".format(i)), os.path.join(loc_polygon, "SVM_{}_polygon.shp".format(i)), "SIMPLIFY", "Classvalue", "SINGLE_OUTER_PART", "")

                    arcpy.AddField_management(os.path.join(loc_polygon, "SVM_{}_polygon.shp".format(i)), "Area", "FLOAT", "", "", "", "", "NULLABLE", "NON_REQUIRED", "")

                    arcpy.CalculateField_management(os.path.join(loc_polygon, "SVM_{}_polygon.shp".format(i)), "Area", "!shape.Area@squaremeters!", "PYTHON_9.3", "")

                    arcpy.MakeFeatureLayer_management(os.path.join(loc_polygon, "SVM_{}_polygon.shp".format(i)), "SVM_{}_selection".format(i), "\"gridcode\" = 1", "", "FID FID VISIBLE NONE;Shape Shape VISIBLE NONE;Id Id VISIBLE NONE;gridcode gridcode VISIBLE NONE;Area Area VISIBLE NONE")

                    arcpy.CalculateField_management("SVM_{}_selection".format(i), "gridcode", "new_class( !Area! )", "PYTHON_9.3", "def new_class(x):\\n    if x <= 10:\\n        return 3\\n    elif x > 10:\\n        return 1\\n")

                    arcpy.AddField_management(os.path.join(loc_polygon, "SVM_{}_polygon.shp".format(i)), "Class_name", "TEXT", "", "", "", "", "NULLABLE", "NON_REQUIRED", "")

                    arcpy.CalculateField_management(os.path.join(loc_polygon, "SVM_{}_polygon.shp".format(i)), "Class_name", "new_class( !gridcode! )", "PYTHON_9.3", "def new_class(x):\\n    if x == 1:\\n        return \"GAP\"\\n    elif x == 2:\\n        return \"WEED\"\\n    else:\\n        return \"SUGARCANE\"\\n\\n")

                    svm = ("svm_{}.tif".format(nama_petak_low)) 
                    svm_shp = ("svm_{}_selection".format(nama_petak_low))
                    petak_polygon = ("SVM_{}_polygon".format(nama_petak))
                    petak_polygon_lowvers = ("svm_{}_polygon".format(nama_petak_low))
                    petak_to_png = ("Petak_{}".format(nama_petak))
                    for df in arcpy.mapping.ListDataFrames(mxd):
                        for lyr in arcpy.mapping.ListLayers(mxd, "", df):
                            if lyr.name.lower() == (svm):
                                arcpy.mapping.RemoveLayer(df, lyr)
                            elif lyr.name.lower() == (svm_shp):
                                arcpy.mapping.RemoveLayer(df, lyr)
                            # elif lyr.name.lower() == (petak_polygon):
                            #     arcpy.mapping.RemoveLayer(df, lyr)
                            # elif lyr.name.lower() == (petak_to_png):
                            #     arcpy.mapping.ExportToPNG(mxd, os.path.join(loc_svm, "gambar.png"))            

                    in_layer = arcpy.mapping.ListLayers(mxd, (petak_polygon), df)[0]
                    in_symbology_layer = arcpy.mapping.Layer(data_symbology[0])
                    arcpy.ApplySymbologyFromLayer_management(in_layer,data_symbology[0])


                    select_feature = arcpy.SelectLayerByAttribute_management(shp, "NEW_SELECTION", "\"Petak_12\" = '{}'".format(i))
                    lyr = arcpy.mapping.ListLayers(mxd, (select_feature), df)[0]
                    df.extent = lyr.getSelectedExtent()
                    # arcpy.df.zoomToSelectedFeatures()
                    arcpy.RefreshActiveView()

                    for df in arcpy.mapping.ListDataFrames(mxd):
                        for lyr in arcpy.mapping.ListLayers(mxd, "", df):
                            if lyr.name.lower() == (petak_polygon_lowvers):
                                lyr.visible = False
                    arcpy.RefreshTOC()
                    arcpy.RefreshActiveView()

                    ptk_keseluruhan = '3ptk.tif' 
                    for df in arcpy.mapping.ListDataFrames(mxd):
                        for lyr2 in arcpy.mapping.ListLayers(mxd, "", df):
                            if df.name.lower() == (ptk_keseluruhan):
                                df.extent = lyr2.getSelectedExtent()
                                # arcpy.RefreshTOC()
                                arcpy.RefreshActiveView()
                    # arcpy.df.zoomToSelectedFeatures()
                    # arcpy.RefreshTOC()
                    arcpy.RefreshActiveView()

                    global loc_dbf_path
                    loc_dbf_path = []
                    loc_dbf_folder = loc_polygon
                    for file in os.listdir(loc_dbf_folder):
                        if file.endswith('.dbf'):
                            loc_dbf_path.append(os.path.join(loc_polygon, file))
                    def statistic_proces():

                        loc_dbf = loc_dbf_path[0]
                        df_bagus = Dbf5(loc_dbf).to_dataframe()

                        kelas = (df_bagus[["Id", "gridcode", "Area", "Class_name"]])
                        df_bagus['Area'] = df_bagus['Area'].round(decimals = 3)
                        gap_area_max = kelas.groupby('Class_name')['Area'].max().to_frame()
                        gap_area_max.columns = ["Area Max (m^2)"]
                        gap_area_min = kelas.groupby('Class_name')['Area'].min().to_frame()
                        gap_area_min.columns = ["Area Min (m^2)"]
                        area_sum = kelas.groupby('Class_name')['Area'].sum().to_frame()
                        result_all = pd.concat([area_sum, gap_area_max, gap_area_min], axis=1, join='inner')
                        result_all2 = pd.concat([area_sum, gap_area_max, gap_area_min], axis=1, join='inner')
                        result_all3 = result_all.assign(AreaHektar=lambda x: (x.Area/10000))
                        result_all2.columns = ['Area (m^2)', 'Area Max (m^2)', 'Area Min (m^2)']
                        list = ['GAP', 'SUGARCANE', 'WEED']
                        result_all['Class'] = list
                        ##                        EXPORT DATA FRAME TO EXCEL
                        result_all2.to_excel(os.path.join(loc_polygon, "result_Tabel.xlsx"))
                        # result_all2.to_excel(os.path.join(dir_svm, "result_Tabel.xlsx")) (INI YANG ORIGINALNYA)
                        # #                       EDITING EXCEL & EXPORT EXCEL EDITED
                        # Call a Workbook() function of openpyxl 
                        # to create a new blank Workbook object
                        wb = openpyxl.load_workbook(os.path.join(loc_polygon, "result_Tabel.xlsx"))
                        # wb = openpyxl.load_workbook(os.path.join(dir_svm, "result_Tabel.xlsx")) (INI YANG ORIGINALNYA)
                        # Get workbook active sheet
                        # from the active attribute. 
                        sheet = wb.active
                        # set the width of the column
                        sheet.column_dimensions['A'].width = 15
                        sheet.column_dimensions['B'].width = 15
                        sheet.column_dimensions['C'].width = 15
                        sheet.column_dimensions['D'].width = 15
                        sheet.row_dimensions[1].height = 25
                        sheet.row_dimensions[2].height = 25
                        sheet.row_dimensions[3].height = 25
                        sheet.row_dimensions[4].height = 25

                        wb.save(os.path.join(loc_polygon,"result_table_edit.xlsx"))

                        # #                       EXCEL TO PNG
                        excel2img.export_img(os.path.join(loc_polygon,"result_table_edit.xlsx"), os.path.join(loc_polygon,"result_table_edit.png"))

                        # #                       MAKING PIE CHART
                        colors = ["#0000FF", "#FFFAF0", "#1C86EE"]
                        explodes = (0.15, 0.05, 0)
                        textprops = {'fontsize':'10'}
                        fig1, ax = plt.subplots()
                        fig1.set_size_inches(4.3,3)
                        plt.subplots_adjust(left=0.0, bottom=0.1, right=0.9)
                        plt.pie(result_all3["AreaHektar"], autopct=lambda p: "{:.1f}%\n ({:.3f} Ha)".format(p, p*sum(result_all3['AreaHektar'])/100), shadow=True, explode=explodes ,colors=colors, startangle=90, textprops=textprops, radius=0.3)
                        plt.legend(labels=result_all["Class"], bbox_to_anchor=(1,0.5), loc="center right", fontsize=8, bbox_transform=plt.gcf().transFigure)
                        plt.gca().axis("equal")
                        # plt.title("Chart of Path", loc='right' ,fontsize=9, bbox={'facecolor':'0.95', 'pad':2})
                        fig1.savefig(os.path.join(loc_polygon, "result_chart.png"))

                    def making_pdf_result():
                        pdf=FPDF(format='letter')
                        pdf.add_page() #always needed
                        pdf.set_font('arial', 'B', 16)
                        pdf.cell(60)
                        pdf.cell(55, 10, 'Result of Process : Field ', 0, 2, 'C')
                        pdf.cell(60)
                        pdf.cell(25, -9.7, (nama_petak), 0, 2, 'C')
                        # pdf.cell(90, 10, '', 0, 2, 'C')
                        pdf.image(os.path.join(loc_polygon,"result_table_edit.png"), x = 52, y = 130, w=110, h=40, type='', link='')
                        pdf.image(os.path.join(loc_polygon, "result_chart.png"), x = 25, y = 170, w=0, h=0, type='', link='')
                        pdf.output(os.path.join(loc_polygon, 'result_.pdf'), 'f')
                    def deleting_files():
                        # Deleting Files
                        os.remove(os.path.join(loc_polygon, "result_Tabel.xlsx"))
                        os.remove(os.path.join(loc_polygon,"result_table_edit.xlsx"))
                        os.remove(os.path.join(loc_polygon,"result_table_edit.png"))
                        os.remove(os.path.join(loc_polygon, "result_chart.png"))

                    statistic_proces()
                    making_pdf_result()
                    deleting_files()
        else:
            messagebox.showinfo("","You don't have avirkey, Please install avirkey first then run the script again!")

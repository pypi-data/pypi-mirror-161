import os

import openpyxl
from loguru import logger
from openpyxl.worksheet.copier import WorksheetCopy


class ExcelReader(object):
    def __init__(self, file_dir, SWCS, output_file, doxygen_dir="./", output_dir="./"):
        self.workbook = openpyxl.load_workbook(file_dir)
        self.doxygen_dir = doxygen_dir
        self.swcs = SWCS
        self.interface = self.workbook["Interface"]
        self.output_dir = output_dir
        self.output_file = os.path.join(output_dir,
                                        output_file if output_file.endswith(".xlsx") else output_file + ".xlsx")
        self._create_fucntion_description_sheet()
        self._create_sheet()

    def _create_fucntion_description_sheet(self):
        for swc in self.swcs:
            target_dict = {"$ModuleName": swc["name"],
                           "$function_dscription": swc.get("function_description", "N/A"),
                           "$cpu_usage": swc.get("cpu_usage", "N/A"),
                           "$memory_usage": swc.get("memory_usage", "N/A")
                           }
            new_ws = self.workbook.create_sheet("SWC description")
            cp = WorksheetCopy(source_worksheet=self.workbook["SWC功能描述"], target_worksheet=new_ws)
            cp.copy_worksheet()
            new_ws.cell(1, 2).value = target_dict["$ModuleName"]
            new_ws.cell(2, 2).value = target_dict["$function_dscription"]
            new_ws.cell(3, 2).value = target_dict["$cpu_usage"]
            new_ws.cell(4, 2).value = target_dict["$memory_usage"]
            self.workbook.remove(self.workbook["SWC功能描述"])

    def _create_sheet(self):
        for swc in self.swcs:
            target_dict = {"$ModuleName$": swc["name"], "$Topic$": swc["topics"], "$DataType$": swc["data_type"],
                           "$HyperLink$": swc["hyper_link"], "$InputSWCs$": swc["input_swcs"],
                           # "$InputValueRange$": swc["input_value_range"],
                           "$Input_RTE_API$": swc["input_rte_api"],
                           "$OutTopic$": swc["out_topics"], "$OutDataType$": swc["out_data_type"],
                           "$OutHyperLink$": swc["out_hyper_link"], "$OutputSWCs$": swc["output_swcs"],
                           # "$OutputValueRange$": swc["output_value_range"],
                           "$Output_RTE_API$": swc["output_rte_api"],
                           }
            new_ws = self.workbook.create_sheet(target_dict["$ModuleName$"])
            cp = WorksheetCopy(source_worksheet=self.interface, target_worksheet=new_ws)
            cp.copy_worksheet()
            n_input_swc = len(swc["input_swcs"])
            n_output_swc = len(swc["output_swcs"])
            if n_input_swc > 1:
                new_ws.insert_rows(10, n_input_swc - 1)
            new_ws.merge_cells("A9:A" + str(9 + max(1, n_input_swc)))
            output_begin_row = 10 + max(1, n_input_swc)
            if n_output_swc > 1:
                new_ws.insert_rows(output_begin_row + 1, n_output_swc - 1)
            new_ws.merge_cells("A{}:A{}".format(output_begin_row, output_begin_row + max(1, n_output_swc)))
            input_columns_of_interest = {2: "$Topic$", 3: "$DataType$", 4: "$HyperLink$",
                                         # 5: "$InputValueRange$",
                                         5: "N/A",
                                         6: "$Input_RTE_API$",
                                         7: "$InputSWCs$"}
            output_columns_of_interest = {2: "$OutTopic$", 3: "$OutDataType$", 4: "$OutHyperLink$",
                                          # 5: "$OutputValueRange$",
                                          5: "N/A",
                                          6: "$Output_RTE_API$",
                                          7: "$OutputSWCs$"}
            input_start = 10
            input_end = 10 + n_input_swc
            output_start = max(12, 10 + n_input_swc + 1)
            output_end = max(12 + n_output_swc, 10 + n_input_swc + 1 + n_output_swc)
            for col in input_columns_of_interest:
                for i in range(input_start, input_end):
                    if col != 4 and col not in [5]:
                        new_ws.cell(i, col).value = target_dict[input_columns_of_interest[col]][i - input_start]
                    elif col in [5]:
                        new_ws.cell(i, col).value = input_columns_of_interest[col]
                    else:
                        doc_file_name = os.path.join(self.doxygen_dir,
                                                     target_dict[input_columns_of_interest[col]][i - input_start])
                        if os.path.exists(doc_file_name):
                            new_ws.cell(i, col).value = '= HYPERLINK（"{}"，"{}"）'.format(
                                doc_file_name,
                                target_dict[input_columns_of_interest[col]][i - input_start].split("_1_1")[-1])
                        else:
                            logger.warning(doc_file_name + ' not exist!')
                            new_ws.cell(i, col).value = 'N/A'
                for j in range(output_start, output_end):
                    if col != 4 and col not in [5]:
                        new_ws.cell(j, col).value = target_dict[output_columns_of_interest[col]][j - output_start]
                    elif col in [5]:
                        new_ws.cell(j, col).value = output_columns_of_interest[col]
                    else:
                        doc_file_name = os.path.join(self.doxygen_dir,
                                                     target_dict[output_columns_of_interest[col]][j - output_start])
                        if os.path.exists(doc_file_name):
                            new_ws.cell(j, col).value = '= HYPERLINK（"{}"，"{}"）'.format(
                                doc_file_name,
                                target_dict[output_columns_of_interest[col]][j - output_start].split("_1_1")[-1])
                        else:
                            logger.warning(doc_file_name + ' not exist!')
                            new_ws.cell(j, col).value = 'N/A'
            for row in new_ws.rows:
                for cell in row:
                    # if cell.value in target_dict:
                    if cell.value == "$ModuleName$":
                        cell.value = target_dict[cell.value] if target_dict[cell.value] else ""
            # add complementary information
            complementary = swc["complementary_msg"]
            new_ws.cell(2, 3).value = complementary.api_period
            new_ws.cell(3, 3).value = complementary.logic_convert_function
            new_ws.cell(6, 3).value = complementary.api_normal_description
            new_ws.cell(6, 6).value = complementary.api_abnormal_description
            new_ws.cell(8, 3).value = complementary.function_work
            new_ws.cell(8, 4).value = complementary.logic_function
            new_ws.cell(8, 5).value = ";".join(complementary.inputs)
            new_ws.cell(8, 6).value = ";".join(complementary.outputs)
            new_ws.cell(8, 7).value = ";".join(complementary.returns)
            new_ws.cell(output_end, 2).value = complementary.last_return

            self.style_excel(new_ws)
        self.workbook.remove(self.interface)
        self.workbook.save(self.output_file)

    def get_num_colnum_dict(self):
        '''
        :return: 返回字典：{1:'A', 2:'B', ...... , 52:'AZ'}
        '''
        num_str_dict = {}
        A_Z = [chr(a) for a in range(ord('A'), ord('Z') + 1)]
        AA_AZ = ['A' + chr(a) for a in range(ord('A'), ord('Z') + 1)]
        A_AZ = A_Z + AA_AZ
        for i in A_AZ:
            num_str_dict[A_AZ.index(i) + 1] = i
        return num_str_dict

    def style_excel(self, sheet):
        '''
        :param sheet_name:  excel中的sheet名
        :return:
        '''
        # 获取最大行数与最大列数
        max_column = sheet.max_column
        max_row = sheet.max_row

        # 将每一列，单元格列宽最大的列宽值存到字典里，key:列的序号从1开始(与字典num_str_dic中的key对应)；value:列宽的值
        max_column_dict = {}

        # 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
        num_str_dict = self.get_num_colnum_dict()

        # 遍历全部列
        for i in range(1, max_column + 1):
            # 遍历每一列的全部行
            for j in range(1, max_row + 1):
                column = 0
                # 获取j行i列的值
                sheet_value = sheet.cell(row=j, column=i).value
                # 通过列表生成式生成字符列表，将当前获取到的单元格的str值的每一个字符放在一个列表中（列表中一个元素是一个字符）
                sheet_value_list = [k for k in str(sheet_value)]
                # 遍历当前单元格的字符列表
                for v in sheet_value_list:
                    # 判定长度，一个数字或一个字母，单元格列宽+=1.1，其它+=2.2（长度可根据需要自行修改，经测试一个字母的列宽长度大概为1）
                    if v.isdigit() == True or v.isalpha() == True:
                        column += 1.1
                    else:
                        column += 2.2
                # 当前单元格列宽与字典中的对比，大于字典中的列宽值则将字典更新。如果字典没有这个key，抛出异常并将值添加到字典中
                try:
                    if column > max_column_dict[i]:
                        max_column_dict[i] = column
                except Exception as e:
                    max_column_dict[i] = column
        # 此时max_column_dict字典中已存有当前sheet的所有列的最大列宽值，直接遍历字典修改列宽
        for key, value in max_column_dict.items():
            sheet.column_dimensions[num_str_dict[key]].width = value

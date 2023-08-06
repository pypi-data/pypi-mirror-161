#需要安装openpyxl库，安装方法为   pip install openpyxl
import pandas as pd
import random
from SAmodel import SAmodel
import numpy as np
import time
random.seed()
import openpyxl
import csv
import matplotlib.pyplot as plt


#将列表中a，b索引之间的所有元素倒序
def reverselist(l, a, b):
    if a > b:
        tmp = a
        a = b
        b = tmp
    u = l[a:b+1:1]
    for i in range(len(u)):
        l[b - i] = u[i]
    return l

class SA(SAmodel):
    def __init__(self, data_path='processed_data.xlsx', NN=200, T0=1, minT=0.005, alpha=0.995, N=1000, epoch=20, maxtest=1500):
        super(SA, self).__init__(T0, minT, alpha, N, epoch, maxtest)
        self.data = pd.read_excel(data_path, 'condition', header=None).values.tolist()
        self.row = len(self.data)
        self.col = len(self.data[0])
        self.stack = None
        self.NN = NN
        #生成一组随机解
        self.xes = self.generate_xes()
        self.E2 = 0.0

    def add(self, x):
        cnt = 0
        for i in range(self.col):
            self.stack[i] = self.stack[i] or x[i]
            cnt += self.stack[i]
        return cnt

    def f(self, x):
        cnt = 1.0
        self.stack = self.data[x[0]].copy()
        for i in x:
            #按照顺序调入订单
            u = self.add(self.data[i])
            #如果当前批次加入此订单会导致货架超标，则开启新的批次并将该订单加入新的批次
            if u > self.NN:
                cnt += 1.0
                self.stack = self.data[i].copy()
        return cnt

    #选取初始解中最好的那几个
    def generate_xes(self, b=50):
        a = [i for i in range(self.row)]
        x = [a.copy() for i in range(b * self.epoch)]
        e = []
        for i in range(b * self.epoch):
            random.shuffle(x[i])
            e.append(self.f(x[i]))
        ids = sorted(range(len(e)), key=lambda k: e[k])
        xx = [x[ids[i]] for i in range(self.epoch)]
        return xx

    #从初始解集合中按顺序选取下一个初始解
    def init_x(self):
        self.pepoch += 1
        return self.xes[self.pepoch-1]

    def Select(self):
        a = random.randint(0, self.row - 1)
        while True:
            b = random.randint(0, self.row - 1)
            if not (b == a):
                break
        self.newx = self.x.copy()
        self.newx = reverselist(self.newx, a, b)
        self.newE = self.f(self.newx)

    def find(self, listpath='processed_data.xlsx'):
        records = []
        for i in range(self.epoch):
            print("epoch " + str(i+1) + " starts")
            time1 = time.perf_counter()
            records.append(self.single_epoch())
            time2 = time.perf_counter()
            print("epoch " + str(i+1) + " finishes with " + str(time2-time1) + "s")
            print("E is " + str(self.E))
        self.generate_answer(listpath)
        return records

    def approach2(self):
        l = [ [] ]
        bag = [ self.data[0].copy() ]
        value = [sum(bag[0])]
        for i in range(self.row):
            minp = -1
            minv = -1.0
            for j in range(len(bag)):
                self.stack = bag[j].copy()
                n = self.add(self.data[i])
                if n <= self.NN and (minp == -1 or minv > n - value[j]):
                    minp = j
                    minv = n - value[j]
            if minp == -1: #没有合适的位置
                l.append([i])
                bag.append(self.data[i].copy())
                value.append(sum(self.data[i]))
            else:  #放入第j个批次中
                j = minp
                l[j].append(i)
                self.stack = bag[j].copy()
                value[j] = self.add(self.data[i])
                bag[j] = self.stack.copy()
        return l

    def generate_answer(self, listpath='processed_data.xlsx'):
        x = self.bestx
        self.stack = self.data[x[0]].copy()
        cnt = 0
        l = [ [] ]
        for i in x:
            # 按照顺序调入订单
            u = self.add(self.data[i])
            # 如果当前批次加入此订单会导致货架超标，则开启新的批次并将该订单加入新的批次
            if u > self.NN:
                cnt += 1
                l.append([])
                self.stack = self.data[i].copy()
            l[cnt].append(i)
        l2 = self.approach2()
        if len(l2) < len(l):
            l = l2
            self.bestE = len(l2)
            cnt = len(l2) - 1
        ll = [min(l[i]) for i in range(cnt+1)]
        ids = sorted(range(cnt+1), key=lambda k: ll[k])
        names = pd.read_excel(listpath, 'orders', header=None).values.tolist()
        for i in range(cnt+1):
            id = ids[i]
            for x in l[id]:
                names[x].append(i+1)
        with open("result1.csv", 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['OrderNo', 'GroupNo'])
            for i in range(len(names)):
                writer.writerow([names[i][0], str(names[i][1])])
        file_name = 'Q2input.xlsx'
        wb = openpyxl.Workbook()
        for i in range(cnt+1):
            id = ids[i]
            ll = l[id]
            sheet = str(i+1)
            sheet2 = 'orders' + str(i+1)
            wb.create_sheet(sheet2)
            ss = wb[sheet2]
            rs = [ll[j] for j in range(len(ll))]
            ss.append(rs)
            wb.create_sheet(sheet)
            ss = wb[sheet]
            for k in ll:
                kk = self.data[k]
                rs = [kk[j] for j in range(len(kk))]
                ss.append(rs)
        wb.save(file_name)


if __name__ == '__main__':
    a = SA(epoch=5, N=800, T0=2, minT=0.01)
    r = a.find()
    print(a.bestE)
    plt.plot(r[a.bestid])
    plt.savefig("Q1.jpg")
    plt.show()

